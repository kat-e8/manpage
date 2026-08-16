"""
Ingest / normalization module.

Reads coder_commands.xlsx (14 sheets, inconsistent shape) and produces a flat
list of rows matching the `commands` table schema (db/schema.sql). This module
only *parses and normalizes* -- it does not touch Postgres. watcher.py imports
`normalize_workbook()` and owns the transactional write.

Sheet irregularities this module was written against (see coderCommands_Plan/
Step03_IngestNormalization.pdf for the full analysis):
  - 6 of 14 sheets have a header row starting with "COMMAND"; 8 do not.
  - docker / ansible / postgres use single-cell rows in ALL CAPS as section
    dividers (e.g. "IMAGE MANAGEMENT") that apply to every following row
    until the next divider.
  - jython uses a two-level outline: column A holds a category ("list",
    "tuple", ...) that is itself a real command row, and is followed by many
    rows with column A blank where the command has shifted into column B.
    These are genuine sub-entries, not prose.
  - shell / git / python / uPython have a handful of blank-column-A rows that
    are NOT shifted sub-entries -- they're free-text continuations of the
    row above (a wrapped explanation). Treating them as their own command
    would fabricate nonsense commands, so they're appended to the previous
    row's `extra` field instead.
  - A sheet is treated as "outline mode" (shift + own sub-entry) only when
    more than 30% of its data rows have a blank column A -- that threshold
    separates jython (75% blank-A) from the sparse stragglers elsewhere
    (git 2%, python 11%, uPython 10%, shell 2%).
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl

HEADER_FIRST_CELL = "COMMAND"
OUTLINE_MODE_THRESHOLD = 0.30


@dataclass
class NormalizedRow:
    topic: str
    section: str | None
    command: str
    description: str | None
    extra: str | None
    row_order: int


def _clean(v) -> str:
    if v is None:
        return ""
    return str(v).replace("\r\n", "\n").strip()


def _section_divider_label(cells: list[str]) -> str | None:
    """A section divider is exactly one populated cell whose text reads like
    a category label (no lowercase letters at all) rather than command
    syntax or prose -- e.g. "IMAGE MANAGEMENT", not "listdir() lists files
    in the root directory of the esp32" (single-cell but real data)."""
    populated = [c for c in cells if c]
    if len(populated) != 1:
        return None
    text = populated[0]
    if len(text) > 60 or text == text.lower():
        return None
    return text if text == text.upper() else None


def _looks_like_header(cells: list[str]) -> bool:
    return bool(cells) and cells[0].strip().upper() == HEADER_FIRST_CELL


def _join_extra(cells: list[str]) -> str | None:
    parts = [c for c in cells if c]
    return " | ".join(parts) if parts else None


def normalize_sheet(sheet_name: str, rows: list[list]) -> list[NormalizedRow]:
    cleaned_rows = [[_clean(c) for c in raw] for raw in rows]
    data_rows = cleaned_rows
    if cleaned_rows and _looks_like_header(cleaned_rows[0]):
        data_rows = cleaned_rows[1:]

    non_blank = [r for r in data_rows if any(r)]
    blank_first = sum(1 for r in non_blank if not r[0] and any(r[1:]))
    outline_mode = bool(non_blank) and (blank_first / len(non_blank)) > OUTLINE_MODE_THRESHOLD

    out: list[NormalizedRow] = []
    current_section: str | None = None   # from an ALL-CAPS divider row
    last_heading: str | None = None      # from the most recent col-A-populated row (outline mode)
    last_row: NormalizedRow | None = None
    order = 0

    for cells in data_rows:
        if not any(cells):
            continue

        divider = _section_divider_label(cells)
        if divider:
            current_section = divider
            continue

        if cells[0]:
            command = cells[0]
            description = cells[1] if len(cells) > 1 and cells[1] else None
            extra = _join_extra(cells[2:])
            order += 1
            row = NormalizedRow(
                topic=sheet_name, section=current_section, command=command,
                description=description, extra=extra, row_order=order,
            )
            out.append(row)
            last_heading = command
            last_row = row
            continue

        # column A is blank
        if outline_mode:
            shifted = cells[1:]
            command = shifted[0] if shifted else ""
            if not command:
                continue
            description = shifted[1] if len(shifted) > 1 and shifted[1] else None
            extra = _join_extra(shifted[2:])
            order += 1
            row = NormalizedRow(
                topic=sheet_name,
                section=current_section or last_heading,
                command=command, description=description, extra=extra,
                row_order=order,
            )
            out.append(row)
            last_row = row
        else:
            # sparse stray row: free-text continuation of the row above
            if last_row is None:
                continue
            text = " ".join(c for c in cells if c)
            last_row.extra = f"{last_row.extra} | {text}" if last_row.extra else text

    return out


def normalize_workbook(xlsx_path: str | Path) -> dict[str, list[NormalizedRow]]:
    """Returns {topic: [NormalizedRow, ...]} for every sheet in the workbook."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    result: dict[str, list[NormalizedRow]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        raw_rows = [list(row) for row in ws.iter_rows(values_only=True)]
        result[sheet_name] = normalize_sheet(sheet_name, raw_rows)
    wb.close()
    return result


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "coder_commands.xlsx"
    data = normalize_workbook(path)
    total = 0
    for topic, rows in data.items():
        print(f"{topic:12s} {len(rows):4d} rows")
        total += len(rows)
    print(f"{'TOTAL':12s} {total:4d} rows")
